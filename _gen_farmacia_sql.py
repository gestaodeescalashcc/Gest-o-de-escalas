"""Parse 6 Junho 2026 - Escala Farmacia HECC.xlsx -> JSON + SQL idempotente.

Padrao igual ao _gen_fisio_sql.py:
  XLSX -> JSON intermediario -> SQL para Supabase MCP.

Departamento alvo: Farmacia
Mes/ano: Junho/2026
"""
import json
import openpyxl

XLSX = r'C:/Users/adoni/OneDrive/Documentos/Gestão de escala/6 Junho 2026 - Escala Farmácia HECC.xlsx'
OUT_JSON = r'C:/Users/adoni/OneDrive/Documentos/Gestão de escala/_farmacia_jun_2026.json'
OUT_SQL = r'C:/Users/adoni/OneDrive/Documentos/Gestão de escala/import_farmacia_junho_2026.sql'

PREP = {'de', 'da', 'das', 'do', 'dos', 'e'}
def title_case(s: str) -> str:
    return ' '.join(w.lower() if w.lower() in PREP else w.capitalize()
                    for w in s.lower().split())

CODE_TO_NAME = {
    'SD': 'Serviço Diurno (7h às 19h) 12h',
    'SN': 'Serviço Noturno (19h às 7h) 12h',
    'MT': 'Manhã e Tarde (8h às 17h) 8h',
    'M':  'Manhã (7h às 13h) 6h',
    'M2': 'Manhã (8h às 12h) 4h',
    'T':  'Tarde (12h às 18h) 6h',
    'P':  'Plantão 24h (7h às 7h) 24h',
}
CODE_TO_TIME = {
    'SD': ('07:00', '19:00'), 'SN': ('19:00', '07:00'),
    'MT': ('08:00', '17:00'), 'M':  ('07:00', '13:00'),
    'M2': ('08:00', '12:00'), 'T':  ('12:00', '18:00'),
    'P':  ('07:00', '07:00'),
}
# Codigos que NAO geram turno (afastamentos): registrados mas ignorados aqui
NON_SHIFT = {'LM', 'LG', 'FE', 'OU', 'FC', 'FD', 'INSS'}

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['CGR']

# Layout: col7 = dia 1 ... col37 = dia 31; profs em rows 7..17
profs = []
for r in range(7, 18):
    nome = ws.cell(row=r, column=2).value
    funcao = ws.cell(row=r, column=3).value
    lotacao = ws.cell(row=r, column=4).value
    ch = ws.cell(row=r, column=5).value
    matricula = ws.cell(row=r, column=1).value
    if not nome:
        continue
    shifts = {}
    for col in range(7, 38):  # dias 1..31
        v = ws.cell(row=r, column=col).value
        if v is None:
            continue
        code = str(v).strip().upper()
        if not code or code in NON_SHIFT:
            continue
        day = col - 6
        shifts[str(day)] = code
    profs.append({
        'nome': nome.strip(),
        'matricula': str(matricula).strip() if matricula not in (None, '') else '',
        'funcao': funcao.strip() if funcao else '',
        'lotacao': lotacao.strip() if lotacao else '',
        'ch': ch.strip() if ch else '',
        'shifts': shifts,
    })

with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(profs, f, ensure_ascii=False, indent=2)
print(f'JSON ok: {len(profs)} profs')

def esc(s: str) -> str:
    return str(s).replace("'", "''")

profs_values = []
shifts_values = []
for p in profs:
    nome_pretty = title_case(p['nome'])
    funcao = p['funcao'].lower()
    if 'atendente' in funcao:
        cat = 'Atendente de Farmácia'
    else:
        cat = 'Farmacêutico'
    profs_values.append(f"('{esc(nome_pretty)}', '{esc(p['matricula'])}', '{esc(cat)}')")
    for day, code in p['shifts'].items():
        full = CODE_TO_NAME.get(code, code)
        st, en = CODE_TO_TIME.get(code, ('00:00', '00:00'))
        shifts_values.append(
            f"('{esc(nome_pretty)}', {day}, '{esc(full)}', '{st}'::time, '{en}'::time)"
        )

profs_v = ',\n  '.join(profs_values)
shifts_v = ',\n  '.join(shifts_values)
total = len(shifts_values)

sql = f"""-- =============================================================================
-- IMPORT FARMACIA -- JUNHO/2026 (planilha "6 Junho 2026 - Escala Farmacia HECC.xlsx")
-- {len(profs)} profissionais, {total} turnos
-- =============================================================================
-- Estrategia:
-- 1) panorama da Farmacia (quem casa com a planilha)
-- 2) limpa escala existente de junho/2026 da Farmacia (shifts/swaps/monthly_schedule)
-- 3a) cria profs novos da planilha (Andressa)
-- 3b) atualiza profs existentes: matricula da planilha sobrescreve, ativa e fixa categoria/departamento
-- 4) NAO desativa profs fora da planilha (Lais Cardoso coord + Simara podem ser ferias/coord)
--    Mas: consolida duplicata "Lais Cardosos dos Anjos" (sem matricula) -> desativa
--    Desativa duplicatas do depto CAF (lixo legado sem matricula)
-- 5) cria a escala de junho/2026 e insere os turnos
-- =============================================================================

CREATE OR REPLACE FUNCTION norm_name(s text) RETURNS text AS $func$
  SELECT lower(translate(s,
    'áàâãäéèêëíìîïóòôõöúùûüçÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ',
    'aaaaaeeeeiiiiooooouuuucAAAAAEEEEIIIIOOOOOUUUUC'));
$func$ LANGUAGE SQL IMMUTABLE;

CREATE TEMP TABLE _planilha_profs (
  nome text PRIMARY KEY,
  matricula text,
  categoria text
);
INSERT INTO _planilha_profs (nome, matricula, categoria) VALUES
  {profs_v};

-- PASSO 1: panorama
SELECT
  p.id, p.full_name, p.registration_number AS matricula, c.name AS categoria, p.active,
  CASE
    WHEN EXISTS(SELECT 1 FROM _planilha_profs pp WHERE norm_name(pp.nome) = norm_name(p.full_name))
    THEN 'na planilha'
    ELSE 'FORA da planilha'
  END AS status_planilha
FROM professionals p
JOIN departments d ON d.id = p.department_id
LEFT JOIN professional_categories c ON c.id = p.category_id
WHERE d.name = 'Farmácia'
ORDER BY status_planilha, p.full_name;

-- PASSO 2: limpa junho/2026 da Farmacia
WITH old_shifts AS (
  SELECT s.id
  FROM shifts s
  JOIN monthly_schedules ms ON ms.id = s.schedule_id
  JOIN departments d ON d.id = ms.department_id
  WHERE d.name = 'Farmácia'
    AND EXTRACT(year FROM ms.month) = 2026
    AND EXTRACT(month FROM ms.month) = 6
)
DELETE FROM shift_swaps
WHERE original_shift_id IN (SELECT id FROM old_shifts)
   OR offered_shift_id IN (SELECT id FROM old_shifts);

DELETE FROM shifts WHERE schedule_id IN (
  SELECT ms.id FROM monthly_schedules ms
  JOIN departments d ON d.id = ms.department_id
  WHERE d.name = 'Farmácia'
    AND EXTRACT(year FROM ms.month) = 2026 AND EXTRACT(month FROM ms.month) = 6
);

DELETE FROM monthly_schedules ms
USING departments d
WHERE ms.department_id = d.id
  AND d.name = 'Farmácia'
  AND EXTRACT(year FROM ms.month) = 2026 AND EXTRACT(month FROM ms.month) = 6;

-- PASSO 3a: cria profs novos (Andressa)
WITH src AS (
  SELECT pp.nome, pp.matricula, pp.categoria,
    (SELECT id FROM professionals p
     JOIN departments d ON d.id = p.department_id
     WHERE norm_name(p.full_name) = norm_name(pp.nome)
       AND d.name = 'Farmácia'
     ORDER BY p.created_at LIMIT 1) AS existing_id
  FROM _planilha_profs pp
)
INSERT INTO professionals (full_name, registration_number, category_id, department_id, company_id, active)
SELECT
  src.nome,
  NULLIF(src.matricula, ''),
  (SELECT id FROM professional_categories WHERE name = src.categoria LIMIT 1),
  (SELECT id FROM departments WHERE name = 'Farmácia' LIMIT 1),
  (SELECT id FROM companies WHERE name ILIKE 'FESF%' LIMIT 1),
  true
FROM src
WHERE src.existing_id IS NULL;

-- PASSO 3b: atualiza profs existentes em Farmacia (sobrescreve matricula, ajusta categoria, ativa)
UPDATE professionals p
SET
  active = true,
  registration_number = CASE
    WHEN NULLIF(pp.matricula, '') IS NOT NULL THEN pp.matricula
    ELSE p.registration_number
  END,
  category_id = (SELECT id FROM professional_categories WHERE name = pp.categoria LIMIT 1),
  updated_at = now()
FROM _planilha_profs pp,
     departments d
WHERE p.department_id = d.id
  AND d.name = 'Farmácia'
  AND norm_name(p.full_name) = norm_name(pp.nome);

-- PASSO 4a: consolida duplicata de Lais Cardoso (a sem matricula vai pra inativo)
UPDATE professionals p
SET active = false, updated_at = now()
FROM departments d
WHERE p.department_id = d.id
  AND d.name = 'Farmácia'
  AND norm_name(p.full_name) LIKE 'lais cardoso%'
  AND p.registration_number IS NULL;

-- PASSO 4b: desativa duplicatas legadas em CAF (sem matricula) cujos nomes batem com a planilha
UPDATE professionals p
SET active = false, updated_at = now()
FROM departments d
WHERE p.department_id = d.id
  AND d.name = 'CAF'
  AND p.registration_number IS NULL
  AND EXISTS (
    SELECT 1 FROM _planilha_profs pp WHERE norm_name(pp.nome) = norm_name(p.full_name)
  );

-- PASSO 5a: cria escala
INSERT INTO monthly_schedules (department_id, name, month, status)
SELECT d.id, 'Escala Farmácia - Junho de 2026', '2026-06-01'::date, 'Rascunho'
FROM departments d
WHERE d.name = 'Farmácia';

-- PASSO 5b: insere turnos
WITH shifts_data(nome, day, shift_type, start_t, end_t) AS (
  VALUES
  {shifts_v}
),
prof_resolved AS (
  SELECT sd.*,
    (SELECT p.id FROM professionals p
     JOIN departments d ON d.id = p.department_id
     WHERE norm_name(p.full_name) = norm_name(sd.nome)
       AND d.name = 'Farmácia'
       AND p.active = true
     ORDER BY p.created_at LIMIT 1) AS prof_id
  FROM shifts_data sd
)
INSERT INTO shifts (professional_id, schedule_id, department_id, shift_date, shift_type, start_time, end_time)
SELECT pr.prof_id, ms.id, d.id, make_date(2026, 6, pr.day), pr.shift_type, pr.start_t, pr.end_t
FROM prof_resolved pr
JOIN departments d ON d.name = 'Farmácia'
JOIN monthly_schedules ms ON ms.department_id = d.id AND ms.month = '2026-06-01'::date
WHERE pr.prof_id IS NOT NULL;

-- VERIFICACAO
SELECT d.name AS setor, COUNT(*) AS total_turnos
FROM shifts s
JOIN monthly_schedules ms ON ms.id = s.schedule_id
JOIN departments d ON d.id = ms.department_id
WHERE EXTRACT(year FROM ms.month) = 2026 AND EXTRACT(month FROM ms.month) = 6
  AND d.name = 'Farmácia'
GROUP BY d.name;

SELECT pp.nome, pp.matricula, 'NAO CASOU' AS aviso
FROM _planilha_profs pp
WHERE NOT EXISTS (
  SELECT 1 FROM professionals p
  JOIN departments d ON d.id = p.department_id
  WHERE norm_name(p.full_name) = norm_name(pp.nome)
    AND d.name = 'Farmácia'
    AND p.active = true
);
"""

with open(OUT_SQL, 'w', encoding='utf-8') as f:
    f.write(sql)
print(f'SQL ok: {total} shifts -> {OUT_SQL}')
