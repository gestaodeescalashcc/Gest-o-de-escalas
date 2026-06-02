"""Export via openpyxl — gera arquivo Excel limpo a partir do template original.
Estratégia: abre o ESCALA MODELO.xlsx ORIGINAL, deixa só "Mai 2026", preenche
dados e salva. Não passa pela exceljs (que corrompe).
"""
import os, sys, re, json, urllib.request
from openpyxl import load_workbook
from copy import copy

SUPABASE_URL = 'https://sbncaocybjiiynktxfqq.supabase.co'
ANON = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InNibmNhb2N5YmppaXlua3R4ZnFxIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzU1ODE2MDQsImV4cCI6MjA5MTE1NzYwNH0.IKKxST21SBc9Zpjj_KRpZFkotP9eXfhiTnlcsrgKQYM'

def login(email, pwd):
    req = urllib.request.Request(
        f'{SUPABASE_URL}/auth/v1/token?grant_type=password',
        data=json.dumps({'email': email, 'password': pwd}).encode(),
        headers={'apikey': ANON, 'Content-Type': 'application/json'},
        method='POST')
    return json.loads(urllib.request.urlopen(req).read())['access_token']

def query(table, params, token):
    url = f'{SUPABASE_URL}/rest/v1/{table}?{params}'
    req = urllib.request.Request(url, headers={
        'apikey': ANON, 'Authorization': f'Bearer {token}',
        'Accept': 'application/json'})
    return json.loads(urllib.request.urlopen(req).read())

token = login('anafarini@fesfsus.ba.gov.br', 'HECC@2026')

SCHED_ID = '67f42d87-c973-4de2-a948-f96d5a197f58'
sched = query('monthly_schedules',
    f'id=eq.{SCHED_ID}&select=*,department:departments!department_id(id,name)', token)[0]
DEPT_ID = sched['department']['id']
profs = query('professionals',
    f'department_id=eq.{DEPT_ID}&active=eq.true&select=id,full_name,registration_number,coren,contracted_hours_per_month,category:professional_categories!category_id(name)&order=full_name', token)
shifts = query('shifts',
    f'schedule_id=eq.{SCHED_ID}&select=professional_id,shift_date,shift_type,deleted_in_realizada_at', token)

# Mapa shift_type → código
TYPES = {
    'Serviço Noturno (19h às 7h) 12h':'SN', 'Serviço Diurno (7h às 19h) 12h':'SD',
    'Plantão 24h (7h às 7h) 24h':'P', 'Manhã (7h às 13h) 6h':'M',
    'Manhã (8h às 12h) 4h':'M2', 'Tarde (12h às 18h) 6h':'T', 'Tarde (13h às 19h) 6h':'T',
    'Manhã e Tarde (8h às 17h) 8h':'MT', 'Folga':'FG', 'Feriado':'FR', 'Férias':'FE',
    'Falta':'FA', 'Licença Prêmio':'LP', 'Licença Médica':'LM', 'Licença Gestação':'LG',
    'Afastamento À Serviço':'AS',
}
HOURS = {'P':24,'SD':12,'SN':12,'MT':8,'M':6,'M2':4,'T':6}
NON_WORK = {'FG','FR','FE','FA','LP','LM','LG','AS'}

# Carrega TEMPLATE ORIGINAL (não o cleaned, pra ter os estilos)
wb = load_workbook('ESCALA MODELO.xlsx')
# Mantém só "Mai 2026"
for name in list(wb.sheetnames):
    if name != 'Mai 2026':
        del wb[name]
ws = wb['Mai 2026']
# Remove imagens (causam crash em alguns leitores)
if hasattr(ws, '_images'): ws._images = []
# Renomeia
ws.title = 'MAIO 26'

# Mês/ano (AH2)
from datetime import date
yy, mm, _ = sched['month'].split('-')
ws['AH2'].value = date(int(yy), int(mm), 1)

# Setor (B4) — uppercase
ws['B4'].value = sched['department']['name'].upper()

# Helper para copiar estilo de uma célula para outra
def copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)

# Linhas modelo (com formatação) no template
TEMPLATE_FIRST_DATA_ROW = 7
TEMPLATE_LAST_DATA_ROW = 26  # template tem ~20 linhas modelo

# Limpa CONTEÚDO das linhas modelo, mantém formatação
for r in range(TEMPLATE_FIRST_DATA_ROW, TEMPLATE_LAST_DATA_ROW + 1):
    for c in range(1, 48):  # A..AU
        cell = ws.cell(row=r, column=c)
        if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
            cell.value = None

# Preenche profissionais. Se exceder TEMPLATE_LAST_DATA_ROW, copia estilo do modelo
firstDayCol = 7  # G
MODEL_ROW = TEMPLATE_FIRST_DATA_ROW  # linha 7 como referência de estilo
NUM_COLS = 47  # vai até coluna AU (totais incluídos)

# Backup da altura padrão da linha modelo
model_height = ws.row_dimensions[MODEL_ROW].height

row = TEMPLATE_FIRST_DATA_ROW
for p in profs:
    # Se passou do limite do template, replica estilo da linha modelo
    if row > TEMPLATE_LAST_DATA_ROW:
        if model_height:
            ws.row_dimensions[row].height = model_height
        for c in range(1, NUM_COLS + 1):
            copy_cell_style(ws.cell(row=MODEL_ROW, column=c), ws.cell(row=row, column=c))

    ws.cell(row=row, column=1).value = p.get('registration_number') or ''
    ws.cell(row=row, column=2).value = p['full_name']
    ws.cell(row=row, column=3).value = (p.get('category') or {}).get('name', '')
    ws.cell(row=row, column=4).value = p.get('coren') or ''
    ws.cell(row=row, column=5).value = p.get('contracted_hours_per_month') or 0
    work_days = 0
    total_h = 0
    day_codes = {}  # day → code (para totalização)
    for s in shifts:
        if s['professional_id'] != p['id']: continue
        if s.get('deleted_in_realizada_at'): continue
        day = int(s['shift_date'][8:10])
        code = TYPES.get(s['shift_type'])
        if not code: continue
        ws.cell(row=row, column=firstDayCol + day - 1).value = code
        if code not in NON_WORK: work_days += 1
        total_h += HOURS.get(code, 0)
        day_codes[day] = code
    ws.cell(row=row, column=6).value = work_days
    ws.cell(row=row, column=40).value = total_h
    row += 1

last_data_row = row - 1  # última linha preenchida

# --- TOTALIZAÇÃO POR DIA / POR CÓDIGO ---
# Adiciona linhas de totalização após o último profissional
# Códigos de interesse (compatível com a visão na tela): SD, SN, MT, P, M, T, M2
SUMMARY_CODES = ['SD', 'SN', 'MT', 'P', 'M', 'T', 'M2']
days_in_month = 31  # MAIO

# Para cada código, conta quantos profs têm aquele código em cada dia
def count_code_per_day(code, day):
    """Conta quantos profs têm <code> no <day>."""
    count = 0
    for p in profs:
        for s in shifts:
            if s['professional_id'] != p['id']: continue
            if s.get('deleted_in_realizada_at'): continue
            if int(s['shift_date'][8:10]) != day: continue
            c = TYPES.get(s['shift_type'])
            if c == code:
                count += 1
                break
    return count

# Estilo das linhas de total (bold + fundo claro)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
total_font = Font(name='Arial', size=10, bold=True)
total_fill = PatternFill('solid', start_color='FFF2F2F2', end_color='FFF2F2F2')
thin = Side(style='thin', color='FF000000')
total_border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')

# Pula uma linha em branco depois do último prof
summary_start = last_data_row + 2

# Remove merges que invadem a área de totalização (linhas summary_start .. summary_start+len(SUMMARY_CODES))
summary_end = summary_start + len(SUMMARY_CODES) - 1
merges_to_remove = []
for mr in ws.merged_cells.ranges:
    if mr.min_row <= summary_end and mr.max_row >= summary_start:
        merges_to_remove.append(str(mr))
for mr in merges_to_remove:
    ws.unmerge_cells(mr)

for i, code in enumerate(SUMMARY_CODES):
    r = summary_start + i
    # Label "TOTAL <code>" merged em A..F
    label = ws.cell(row=r, column=2)
    label.value = f'TOTAL {code}'
    label.font = total_font
    label.fill = total_fill
    label.border = total_border
    label.alignment = Alignment(horizontal='right', vertical='center')
    # Aplica fill/border nas células de label adjacentes (A, C, D, E, F)
    for c in [1, 3, 4, 5, 6]:
        cell = ws.cell(row=r, column=c)
        cell.fill = total_fill
        cell.border = total_border
    # Conta por dia
    for day in range(1, days_in_month + 1):
        cnt = count_code_per_day(code, day)
        cell = ws.cell(row=r, column=firstDayCol + day - 1)
        cell.value = cnt if cnt > 0 else None
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        cell.alignment = center

tmp_out = '_tmp_openpyxl.xlsx'
wb.save(tmp_out)

# Pós-processa: remove externalLinks, definedNames quebrados e referências [1]/#REF!
import zipfile, shutil
out = 'Escala_Enfermeiros_-_Maio_de_2026.xlsx'
out_v2 = 'Escala_Enfermeiros_-_Maio_de_2026_v2.xlsx'

def cleanup(src, dst):
    with zipfile.ZipFile(src) as zi, zipfile.ZipFile(dst, 'w', zipfile.ZIP_DEFLATED) as zo:
        for n in zi.namelist():
            if 'externalLink' in n: continue
            c = zi.read(n)
            if n.endswith('.xml') or n.endswith('.rels'):
                t = c.decode('utf-8')
                # Remove relacionamentos externalLink
                t = re.sub(r'<Relationship[^>]*externalLink[^>]*/>', '', t)
                # Remove externalReferences nodes
                t = re.sub(r'<externalReferences>.*?</externalReferences>', '', t, flags=re.DOTALL)
                t = re.sub(r'<externalReferences\s*/>', '', t)
                # Remove override content type pra externalLink
                t = re.sub(r'<Override[^>]*externalLink[^>]*/>', '', t)
                # Remove VML declaration sem arquivo .vml
                t = re.sub(r'<Default Extension="vml" ContentType="[^"]*"\s*/>', '', t)
                # Remove definedNames problemáticos (com [1] ou #REF!)
                def keep_safe(m):
                    blk = m.group(0)
                    kept = re.findall(r'<definedName[^>]*name="_xlnm\.[^"]*"[^>]*>[^<]*</definedName>', blk)
                    if not kept: return ''
                    return '<definedNames>' + ''.join(kept) + '</definedNames>'
                t = re.sub(r'<definedNames>.*?</definedNames>', keep_safe, t, flags=re.DOTALL)
                c = t.encode('utf-8')
            zo.writestr(n, c)

try:
    cleanup(tmp_out, out)
    final = out
except PermissionError:
    cleanup(tmp_out, out_v2)
    final = out_v2
os.remove(tmp_out)

print(f'Salvo: {final} ({os.path.getsize(final)} bytes, {len(profs)} profs, {len(shifts)} shifts)')

# Valida
with zipfile.ZipFile(final) as z:
    wb_xml = z.read('xl/workbook.xml').decode('utf-8')
    ct = z.read('[Content_Types].xml').decode('utf-8')
    print(f'  workbook tem [1]: {"[1]" in wb_xml}')
    print(f'  workbook tem #REF!: {"#REF!" in wb_xml}')
    print(f'  workbook tem externalLink: {"externalLink" in wb_xml}')
    print(f'  CT tem vml: {"vml" in ct.lower()}')
